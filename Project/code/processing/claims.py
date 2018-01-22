#!/usr/bin/python
import datetime
import os, sys
from processing import Processor
import pandas as pd
import csv
import openpyxl
import xlrd
from cleanco import cleanco

class ClaimsProcessor(Processor):
    input = "claims"
    out = "claims"
    claimType = "Property Damage"
    fields = ["date","month", "airline", "item", "claim_amount"]
    name = "claims"
    def process(self, in_, out):
        """

        :param in_:
        :param out: csv.DictWriter
        :return:
        """
        book = xlrd.open_workbook(in_)
        sheet = book.sheet_by_index(0)
        for row_index in range(1,sheet.nrows):
            # A B C D E F G H I J K  L  M  N  O  P
            # 0 1 2 3 4 5 6 7 8 9 10 11 12 13 14 15
            # print(sheet.cell(row_index,1).value)
            # print(sheet.row(row_index))
            # print(type(sheet.cell(row_index,1).value))
            try:
                val = sheet.cell(row_index, 1).value if sheet.cell(row_index, 1).value is not "" else sheet.cell(
                    row_index, 2).value
                date = datetime.datetime(*xlrd.xldate_as_tuple(val, book.datemode))
                airline_name = sheet.cell(row_index, 5).value
                val = sheet.cell(row_index, 9).value
                r = {"date": date.date().strftime("%Y-%m-%d"), "month":date.date().strftime("%Y-%m"), "airline": cleanco(airline_name).clean_name(),
                     "item": sheet.cell(row_index, 8).value, "claim_amount": val if val is not "-" else 0}

                l = [sheet.cell(row_index, 2), sheet.cell(row_index, 5), sheet.cell(row_index, 8),
                     sheet.cell(row_index, 9)]
                if not "" in r.values() and all(r.values()):
                    out.writerow(r)
            except Exception as e:
                pass
                # print(e)
                # print("something went wrong with row")
                # print(sheet.row(row_index))
        # wb = openpyxl.load_workbook(in_)
        # ws = wb.get_active_sheet()
        # for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        #
        #     out.writerow({"Incident Date":row["C"],"Airline Name":row["F"],"Item":row["I"],"Claim Amount":row["J"]})

        # xl = pd.ExcelFile(in_)
        # df = pd.DataFrame(xl.parse(xl.sheet_names[0]), columns=self.fields)
        # df.empty
        # out.writer.writerows( df.to_csv().split("\n"))
        # df.to_csv(out)
